import streamlit as st
from bs4 import BeautifulSoup
import pandas as pd
import io
import re
from supabase import create_client, Client
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import time

# 🔹 إعدادات Supabase
SUPABASE_URL = "https://khrqitmudxijsuorwhvj.supabase"
SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImtocnFpdG11ZHhpanN1b3J3aHZqIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NjMwNDcwNDMsImV4cCI6MjA3ODYyMzA0M30.Jb61EjkNFqx7mOz5tX_3yKT-fnX7zBJNyAEVSnpSmxg"
BUCKET_NAME = "uploads"

supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

st.title("📤 رفع ملف ASPX الطالب")

# 🧍 إدخال يدوي للبيانات
student_name = st.text_input("👤 اسم الطالب الكامل")
student_id_input = st.text_input("🎓 الرقم الجامعي")

uploaded_file = st.file_uploader("اختر ملف ASPX", type=["aspx"])

if uploaded_file is not None and student_name and student_id_input:
    try:
        # 🧩 فحص هل الملف موجود مسبقًا
        safe_id = re.sub(r'[^A-Za-z0-9_-]', '_', student_id_input)
        file_name = f"{safe_id}.xlsx"

        existing_files = supabase.storage.from_(BUCKET_NAME).list()
        file_exists = any(f["name"] == file_name for f in existing_files)

        if file_exists:
            st.error(f"⚠️ هذا الرقم الجامعي ({student_id_input}) موجود مسبقًا، لا يمكن رفع الملف مرة أخرى.")
        else:
            # 🧠 قراءة محتوى الملف
            content = uploaded_file.read().decode("utf-8")
            soup = BeautifulSoup(content, "html.parser")
            tables = soup.find_all("table")

            # استخراج بيانات من الصفحة
            full_text = soup.get_text(separator="\n")
            major_match = re.search(r"التخصص\s*[:\-]?\s*(.+)", full_text)
            admission_year_match = re.search(r"سنة القبول\s*[:\-]?\s*(\d{4})", full_text)
            admission_type_match = re.search(r"نوع القبول\s*[:\-]?\s*(.+)", full_text)

            major = major_match.group(1).strip() if major_match else ""
            admission_year = admission_year_match.group(1).strip() if admission_year_match else ""
            admission_type = admission_type_match.group(1).strip() if admission_type_match else ""

            if admission_year:
                start_year = int(admission_year)
                admission_year_full = f"{start_year}/{start_year + 1}"
            else:
                admission_year_full = ""

            all_rows = []

            # معالجة الجداول
            for table in tables:
                title_td = table.find("td", colspan=True)
                if title_td:
                    if all_rows:
                        all_rows.append([""] * 5)
                    continue

                for i, tr in enumerate(table.find_all("tr")):
                    if tr.find("td", colspan=True):
                        continue
                    cells = [td.get_text(strip=True) for td in tr.find_all(["td", "th"])]
                    if not cells:
                        continue
                    if i == 0:
                        row = [student_name, student_id_input, major, admission_year_full, admission_type] + cells
                    else:
                        row = ["", "", "", "", ""] + cells
                    all_rows.append(row)

            if not all_rows:
                st.warning("⚠️ لا توجد بيانات صالحة في الملف.")
            else:
                # ضبط الأعمدة
                max_cols = max(len(r) for r in all_rows)
                for r in all_rows:
                    while len(r) < max_cols:
                        r.append("")
                columns = ["Student Name", "Student ID", "Major", "Admission Year", "Admission Type"] + [
                    f"Column{i}" for i in range(1, max_cols - 5 + 1)
                ]
                df = pd.DataFrame(all_rows, columns=columns)

                # حفظ مؤقت
                excel_buffer = io.BytesIO()
                df.to_excel(excel_buffer, index=False)
                excel_buffer.seek(0)

                # تعديل عرض الأعمدة
                wb = load_workbook(excel_buffer)
                ws = wb.active
                for col in ws.columns:
                    max_length = 0
                    col_letter = get_column_letter(col[0].column)
                    for cell in col:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    ws.column_dimensions[col_letter].width = max_length + 5
                excel_buffer2 = io.BytesIO()
                wb.save(excel_buffer2)
                excel_buffer2.seek(0)

                # رفع إلى Supabase
                res = supabase.storage.from_(BUCKET_NAME).upload(
                    file_name,
                    excel_buffer2.getvalue(),
                    {"content-type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
                )

                if "error" in str(res).lower():
                    st.error(f"❌ حدث خطأ أثناء الرفع: {res}")
                else:
                    st.success(f"✅ تم رفع الملف بنجاح باسم ({file_name})!")

    except Exception as e:
        st.error(f"❌ حدث خطأ أثناء المعالجة: {e}")

elif uploaded_file and (not student_name or not student_id_input):
    st.warning("⚠️ يرجى إدخال الاسم الكامل والرقم الجامعي قبل رفع الملف.")
